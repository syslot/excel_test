#! coding=utf-8

from openpyxl import Workbook,load_workbook
import psycopg2
import datetime,time
import subprocess

from openpyxl.worksheet.filters import AutoFilter

tmp_data_file = '/Users/syslot/Desktop/tmp.xlsx'
tmp_review_file = '/Users/syslot/Desktop/review.xlsx'

try:
    conn = psycopg2.connect("dbname='ielts' user='syslot' password='123456'")
except:
    print "I am unable to connect to the database"

def insert_data(chapter, word_list, datastr = None,  style = 'test'):
    cur = conn.cursor()
    if datastr == None:
        datastr = datetime.datetime.fromtimestamp(time.time()).strftime('%Y-%m-%d')
    tmp = 0
    for word in word_list:
        tmp += 1
        if style == 'formal':
            cur.execute("INSERT INTO origin_data(word, datetime, chapter, oset) VALUES (%s, %s, %s, %s)" , (word, datastr, chapter, tmp))
        else:
            cur.execute("INSERT INTO test_data(word, datetime, chapter, oset) VALUES (%s, %s, %s, %s)", (word, datastr, chapter, tmp))

    rst = conn.commit()
    return rst

def get_count(chapter):
    cur = conn.cursor()
    cur.execute("SELECT count from origin_count WHERE chapter = '%s'" % (chapter))
    rst = cur.fetchall()
    return rst[0][0]

def read_data_from_excel(data_sheet, offset = 0):
    wb = load_workbook(data_sheet)
    tds = {}
    for s_n in wb.sheetnames:
        sht = wb.get_sheet_by_name(s_n)
        row = get_count(s_n)
        list =[ x[0].value for x in sht['A%d:%d' % (1+offset, row +1)]]
        tds[sht.title] = list
    return tds

def insert_origin_data(origin_data_sheet):
    tds = read_data_from_excel(origin_data_sheet, 1)
    for k,v in tds.items():
        insert_data(k, v, style='formal')

def insert_test_data(test_data_sheet, test_date=None):
    tds = read_data_from_excel(test_data_sheet)
    for k,v in tds.items():
        t = None
        if test_date == None:
            t = datetime.datetime.fromtimestamp(time.time()).strftime('%Y-%m-%d')
        else:
            t = test_date
        insert_data(k,v,t)

def get_chp_date(chp):
    cur = conn.cursor()
    cur.execute("SELECT DISTINCT datetime from test_data WHERE chapter = '%s' ORDER BY datetime DESC"%chp)
    l = []
    rst = cur.fetchall()
    return rst[0][0].strftime('%Y-%m-%d')

def get_test_data(chp, dt):
    cur = conn.cursor()
    cur.execute("SELECT word FROM test_data WHERE chapter = %s AND datetime = %s ORDER BY oset ASC" , (chp,dt))
    l = []
    rst = cur.fetchall()
    [l.append(x[0]) if x[0] != None else l.append("") for x in rst]
    return l

def get_formal_data(chp):
    cur = conn.cursor()
    cur.execute("SELECT word FROM origin_data WHERE chapter = '%s' ORDER BY oset"%chp)
    l = []
    rst = cur.fetchall()
    [l.append(x[0]) for x in rst]
    return l

def comp(correct, test):
    rst = []
    t_l = len(test)
    for i in range(t_l):
        if i == "":
            rst.append('W')
            continue
        c_data = correct[i].lower().strip()
        t_data = test[i].lower().strip()
        if c_data != t_data:
            rst.append('W')
        else:
            rst.append('R')

    if (t_l < len(correct)):
        [rst.append("W") for i in range(len(correct) - t_l)]
    return rst

def down_data(test_data_sheet):
    tds = read_data_from_excel(test_data_sheet)
    correct_word = {}
    wb = Workbook()
    for k in tds.keys():
        sht = wb.create_sheet(k)
        correct_word[k] = get_formal_data(k)
        v, width = transpose(correct_word[k], tds[k])
        sht.append(["word","status","check"])
        [sht.append(row) for row in v]
        sht.page_setup.fitToWidth = 1
        sht.column_dimensions["A"].width = width
        sht.column_dimensions["C"].width = width
        af = AutoFilter("A1:C%d" % len(v))
        af.add_filter_column(1, ['\u2715'])
        sht.auto_filter = af
        sht.sheet_view.zoomScale = 200
    del wb['Sheet']
    wb.save(tmp_data_file)

def transpose(correct, test, chp = None, offset = 0):
    view_list = []
    max_size = -1;
    for i in range(len(correct)):
        row = [correct[i],u'=IF(A%d=C%d,"","\u2715")' % (offset + i+2, offset + i+2), test[i]]
        if chp != None:
            row.append(chp)
        view_list.append(row)
        max_size = len(correct[i]) if len(correct[i]) > max_size else max_size
    return view_list, max_size

def gene_tuple(x,y):
    i = 0
    while i < len(x):
        word_pair = (x[i],y[i])
        yield word_pair
        i += 1

def read_chp(chp, delay = 5):
    td = get_test_data(chp, get_chp_date(chp))
    fd = get_formal_data(chp)
    rst = filter(lambda x: x[0].lower() != x[1].lower(), gene_tuple(fd, td))
    print chp, len(rst)
    for t in rst:
        print "origin: %s" % t[0]
        print "error : %s" % t[1]
        subprocess.call('say %s' % t[0].replace('\'', '\\\''), shell=True)
        if ' ' in t[0]:
            time.sleep(delay)
        else:
            time.sleep(delay/2)

def get_review_by_chp(chp):
    last_date = get_chp_date(chp)
    cor_data = get_formal_data(chp)
    test_data = get_test_data(chp, last_date)
    rst = comp(cor_data, test_data)
    c_x = []
    t_x = []
    for i in range(len(test_data)):
        if rst[i]=='W':
            c_x.append(cor_data[i])
            t_x.append(test_data[i])
    if len(cor_data) > len(test_data):
        for word in c_x[len(test_data):]:
            c_x.append(word)
            t_x.append('')
    return c_x, t_x

def get_review(chp_list):
    c_x, t_x =[], []
    wb = Workbook()
    sht = wb.create_sheet("review")
    sht.append(["word", "status", "check", "chapter"])
    max_size = -1
    sum = 0
    for chp in chp_list:
        c, t = get_review_by_chp(chp)
        rows, size = transpose(c, t, chp, sum)
        sum += len(rows)
        max_size = size if size > max_size else max_size
        [sht.append(row) for row in rows]

    sht.page_setup.fitToWidth = 1
    sht.column_dimensions["A"].width = max_size
    sht.column_dimensions["C"].width = max_size
    af = AutoFilter("A1:C%d" % sum)
    af.add_filter_column(1, ['\u2715'])
    sht.auto_filter = af
    sht.sheet_view.zoomScale = 200
    sht.cell('E1').value = "sum"
    sht.cell('E2').value = sum
    del wb['Sheet']
    wb.save(tmp_review_file)

def read_review_file(delay = 5, offset = 0):
    wb = load_workbook(tmp_review_file)
    tds = {}
    sht = wb['review']
    row = sht.cell('e2').value
    list =[ x[0].value for x in sht['A%d:%d' % (2, row +1)]]
    time.sleep(delay/2)
    for t in list[offset:]:
        print t
        subprocess.call('say %s' % t.replace('\'', '\\\''), shell=True)
        if ' ' in t[0]:
            time.sleep(delay)
        else:
            time.sleep(delay/2)
    return list

if __name__ == '__main__':
    # insert_test_data('/Users/syslot/Desktop/listening origin/5-19.xlsx', '2017-05-19')
    # insert_test_data('/Users/syslot/Desktop/7-25.xlsx')
    # insert_origin_data('/Users/syslot/Desktop/2017.xlsx')

    # chp_list=['3-1',
    #           '3-2',
    #           '3-3',
    #           '3-4',
    #           '3-5',
    #           '3-6',
    #           '3-7',
    #           '3-8',
    #           '3-9',
              # '4-1']
    # for chp in chp_list:
    #     read_chp(chp)
    # read_chp('3-5')

    # down_data('/Users/syslot/Desktop/listening origin/5-19.xlsx')

    # down_data('/Users/syslot/Desktop/7-25.xlsx')
    chaps_3 = ['3-1', '3-2', '3-3', '3-4', '3-5', '3-6', '3-7','3-8','3-9']
    chaps_4 = ['4-1', '4-2', '4-3', '4-4']
    chaps_5 = ['5-1', '5-2', '5-3', '5-4', '5-5', '5-6', '5-7','5-8', '5-9', '5-10', '5-11', '5-12']
    get_review(chaps_3 + chaps_4 + chaps_5)
    # read_review_file()
    # read_review_file(8,130)
